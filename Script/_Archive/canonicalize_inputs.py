"""
Canonicalize cargo-type naming in input Excel files (in place).

Goal: Replace spaced variants (e.g., "Break Bulk") with canonical hyphenated names
("Break-Bulk") across relevant inputs so downstream processing can assume canonical
labels and fail fast otherwise.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).parent.parent
PROCESSED_DATA_DIR = BASE_DIR / "Processed_Data"

PIER_INPUT_PATH = PROCESSED_DATA_DIR / "Honolulu Harbor Pier Operations and Cargo Inventory.xlsx"


REPLACEMENTS = {
    "Break Bulk": "Break-Bulk",
    "Dry Bulk": "Dry-Bulk",
    "Liquid Bulk": "Liquid-Bulk",
}


def canonicalize_text(value: str) -> str:
    out = value
    for src, dst in REPLACEMENTS.items():
        out = out.replace(src, dst)
    return out


def canonicalize_pier_workbook(path: Path) -> None:
    wb = load_workbook(path)
    if "Cargo_Piers" not in wb.sheetnames:
        raise ValueError(f"Expected sheet 'Cargo_Piers' in {path.name}. Found: {wb.sheetnames}")

    ws = wb["Cargo_Piers"]

    # Canonicalize header row (row 1)
    header_changes = 0
    for cell in ws[1]:
        if isinstance(cell.value, str):
            new_value = canonicalize_text(cell.value)
            if new_value != cell.value:
                cell.value = new_value
                header_changes += 1

    # Locate the 'Cargo Types' column (after any header edits)
    headers = [c.value for c in ws[1]]
    try:
        cargo_types_col_idx = headers.index("Cargo Types") + 1  # 1-based indexing
    except ValueError:
        cargo_types_col_idx = None

    # Canonicalize 'Cargo Types' cell text (free-form)
    cell_changes = 0
    if cargo_types_col_idx is not None:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=cargo_types_col_idx)
            if isinstance(cell.value, str) and cell.value.strip():
                new_value = canonicalize_text(cell.value)
                if new_value != cell.value:
                    cell.value = new_value
                    cell_changes += 1

    wb.save(path)

    print(f"Updated {path.name} in place.")
    print(f"  - Header cells changed: {header_changes}")
    if cargo_types_col_idx is None:
        print("  - Note: 'Cargo Types' column not found; no cell text canonicalization applied.")
    else:
        print(f"  - 'Cargo Types' cells changed: {cell_changes}")


def main() -> None:
    canonicalize_pier_workbook(PIER_INPUT_PATH)


if __name__ == "__main__":
    main()

